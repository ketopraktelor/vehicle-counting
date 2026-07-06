
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.chart import (
    BarChart,
    Reference
)

# ==========================================================
# KONFIGURASI
# ==========================================================
OUTPUT_FOLDER = "hasil-counting"

CSV_JAM_LANCAR = os.path.join(
    OUTPUT_FOLDER,
    "jam-lancar.csv"
)

CSV_JAM_SIBUK = os.path.join(
    OUTPUT_FOLDER,
    "jam-sibuk.csv"
)

OUTPUT_EXCEL = os.path.join(
    OUTPUT_FOLDER,
    "laporan_lalu_lintas.xlsx"
)

# ==========================================================
# MEMBACA FILE CSV
# ==========================================================

def baca_csv(path):
    if os.path.exists(path):
        print(f"[OK] Membaca : {path}")
        
        return pd.read_csv(path)
    print(f"[INFO] File tidak ditemukan : {path}")

    return None

# ==========================================================
# STYLE WORKSHEET
# ==========================================================
def style_sheet(ws):
    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    total_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin = Side(style="thin")
    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    for column_cells in ws.columns:
        panjang = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )

    if column_cells[0].column_letter == "A":
        ws.column_dimensions["A"].width = 38

    elif column_cells[0].column_letter == "B":
        ws.column_dimensions["B"].width = 45

    else:
        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = panjang + 5
        
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).lower() == "total":
            for cell in row:
                cell.fill = total_fill
                cell.font = Font(bold=True)
                
# ==========================================================
# MEMBUAT GRAFIK BATANG
# ==========================================================

def buat_grafik(worksheet, judul):
    """
    Menambahkan grafik batang ke worksheet.
    Baris 'Total' tidak dimasukkan ke dalam grafik.
    """
    chart = BarChart()
    total_row = worksheet.max_row

    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == "Total":
            total_row = row - 1
            break
        
    data = Reference(
        worksheet,
        min_col=2,
        max_col=2,
        min_row=1,
        max_row=total_row
    )

    kategori = Reference(
        worksheet,
        min_col=1,
        min_row=2,
        max_row=total_row
    )

    chart.add_data(
        data,
        titles_from_data=True
    )

    chart.set_categories(kategori)
    chart.title = judul

    chart.y_axis.title = "Jumlah Kendaraan"
    chart.x_axis.title = "Jenis Kendaraan"

    chart.height = 8
    chart.width = 14

    worksheet.add_chart(
        chart,
        "D2"
    )
                
# ==========================================================
# MEMBUAT SHEET
# ==========================================================
def buat_sheet(workbook, nama_sheet, dataframe):
    """
    Membuat worksheet berdasarkan DataFrame.
    """
    worksheet = workbook.create_sheet(
        title=nama_sheet
    )
    
    worksheet.append(list(dataframe.columns))
    
    for row in dataframe.itertuples(index=False):
        worksheet.append(list(row))

    style_sheet(worksheet)
    buat_grafik(
        worksheet,
        nama_sheet
    )

    return worksheet

# ==========================================================
# MEMBUAT SHEET PERBANDINGAN
# ==========================================================
def buat_sheet_perbandingan(workbook, df_jam_lancar, df_jam_sibuk):
    """
    Membuat worksheet perbandingan antara
    jam lancar dan jam sibuk.
    """
    jam_lancar = df_jam_lancar.rename(
        columns={
            "Jumlah": "Jam Lancar"
        }
    )

    jam_sibuk = df_jam_sibuk.rename(
        columns={
            "Jumlah": "Jam Sibuk"
        }
    )

    df_perbandingan = pd.merge(
        jam_lancar,
        jam_sibuk,
        on="Jenis Kendaraan",
        how="outer"
    ).fillna(0)
    
    urutan = [
        "Car",
        "Motorcycle",
        "Bus",
        "Truck",
        "Total"
    ]
    
    df_perbandingan["Jenis Kendaraan"] = pd.Categorical(
        df_perbandingan["Jenis Kendaraan"],
        categories=urutan,
        ordered=True
    )
    
    df_perbandingan = df_perbandingan.sort_values(
        "Jenis Kendaraan"
    ).reset_index(drop=True)

    worksheet = workbook.create_sheet(
        title="Perbandingan"
    )

    worksheet.append(
        list(df_perbandingan.columns)
    )

    for row in df_perbandingan.itertuples(index=False):
        worksheet.append(list(row))
    style_sheet(worksheet)
    
    return worksheet

# ==========================================================
# MEMBUAT SHEET RINGKASAN
# ==========================================================
def buat_sheet_ringkasan(workbook, df_jam_lancar, df_jam_sibuk):
    """
    Membuat worksheet ringkasan hasil perbandingan.
    """

    worksheet = workbook.create_sheet(title="Ringkasan")

    # =====================================================
    # HEADER TABEL
    # =====================================================
    worksheet.append(["Keterangan", "Nilai"])

    # =====================================================
    # TOTAL KENDARAAN
    # =====================================================
    total_lancar = int(
        df_jam_lancar.loc[
            df_jam_lancar["Jenis Kendaraan"] == "Total",
            "Jumlah"
        ].values[0]
    )

    total_sibuk = int(
        df_jam_sibuk.loc[
            df_jam_sibuk["Jenis Kendaraan"] == "Total",
            "Jumlah"
        ].values[0]
    )

    # =====================================================
    # SELISIH DAN PERSENTASE
    # =====================================================
    selisih = total_sibuk - total_lancar

    if total_lancar != 0:
        persentase = (selisih / total_lancar) * 100
    else:
        persentase = 0

    # =====================================================
    # KENDARAAN TERBANYAK
    # =====================================================
    lancar = df_jam_lancar[
        df_jam_lancar["Jenis Kendaraan"] != "Total"
    ]

    sibuk = df_jam_sibuk[
        df_jam_sibuk["Jenis Kendaraan"] != "Total"
    ]

    kendaraan_lancar = lancar.loc[
        lancar["Jumlah"].idxmax(),
        "Jenis Kendaraan"
    ]

    kendaraan_sibuk = sibuk.loc[
        sibuk["Jumlah"].idxmax(),
        "Jenis Kendaraan"
    ]

    # =====================================================
    # ISI TABEL
    # =====================================================
    worksheet.append([
        "Total Kendaraan Jam Lancar",
        total_lancar
    ])

    worksheet.append([
        "Total Kendaraan Jam Sibuk",
        total_sibuk
    ])

    worksheet.append([
        "Selisih Kendaraan",
        abs(selisih)
    ])

    worksheet.append([
        "Persentase Perubahan",
        f"{abs(persentase):.2f}%"
    ])

    worksheet.append([
        "Kendaraan Terbanyak (Jam Lancar)",
        kendaraan_lancar
    ])

    worksheet.append([
        "Kendaraan Terbanyak (Jam Sibuk)",
        kendaraan_sibuk
    ])

    # =====================================================
    # STYLE TABEL PERTAMA
    # =====================================================
    style_sheet(worksheet)

    # =====================================================
    # MEMBUAT KESIMPULAN
    # =====================================================
    if total_lancar > total_sibuk:

        kesimpulan = (
            f"Berdasarkan hasil perbandingan, jumlah kendaraan pada kondisi jam "
            f"lancar adalah {total_lancar} kendaraan, sedangkan pada kondisi jam "
            f"sibuk adalah {total_sibuk} kendaraan. Selisih jumlah kendaraan "
            f"antara kedua kondisi tersebut adalah {abs(selisih)} kendaraan. "
            f"Dengan demikian, kondisi jam lancar memiliki jumlah kendaraan "
            f"yang lebih tinggi dibandingkan kondisi jam sibuk."
        )

    elif total_sibuk > total_lancar:

        kesimpulan = (
            f"Berdasarkan hasil perbandingan, jumlah kendaraan pada kondisi jam "
            f"sibuk adalah {total_sibuk} kendaraan, sedangkan pada kondisi jam "
            f"lancar adalah {total_lancar} kendaraan. Selisih jumlah kendaraan "
            f"antara kedua kondisi tersebut adalah {abs(selisih)} kendaraan. "
            f"Dengan demikian, kondisi jam sibuk memiliki jumlah kendaraan "
            f"yang lebih tinggi dibandingkan kondisi jam lancar."
        )

    else:

        kesimpulan = (
            f"Berdasarkan hasil perbandingan, jumlah kendaraan pada kondisi jam "
            f"lancar dan jam sibuk sama, yaitu {total_lancar} kendaraan."
        )

    # =====================================================
    # TABEL KESIMPULAN
    # =====================================================
    worksheet.append([])
    worksheet.append(["KESIMPULAN"])
    worksheet.append([kesimpulan])

    worksheet.merge_cells("A9:B9")
    worksheet.merge_cells("A10:B10")

    worksheet["A9"].fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    worksheet["A9"].font = Font(
        bold=True,
        color="FFFFFF"
    )

    worksheet["A9"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet["A10"].alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True
    )

    worksheet.row_dimensions[10].height = 80

# ==========================================================
# PROGRAM UTAMA
# ==========================================================
def main():
    print("=" * 50)
    print("MEMBUAT LAPORAN LALU LINTAS")
    print("=" * 50)

    # ---------------------------------------
    # Membaca CSV
    # ---------------------------------------
    df_jam_lancar = baca_csv(CSV_JAM_LANCAR)
    if df_jam_lancar is None:
        print("\nFile jam-lancar.csv tidak ditemukan!")
        return
    df_jam_sibuk = baca_csv(CSV_JAM_SIBUK)

    # ---------------------------------------
    # Membuat Workbook
    # ---------------------------------------
    workbook = Workbook()
    workbook.remove(workbook.active)

    # ---------------------------------------
    # Sheet Jam Lancar
    # ---------------------------------------
    buat_sheet(
        workbook,
        "Jam Lancar",
        df_jam_lancar
    )

    # ---------------------------------------
    # Sheet Jam Sibuk
    # ---------------------------------------
    if df_jam_sibuk is not None:

        buat_sheet(
        workbook,
        "Jam Sibuk",
        df_jam_sibuk
    )

    buat_sheet_perbandingan(
        workbook,
        df_jam_lancar,
        df_jam_sibuk
    )

    buat_sheet_ringkasan(
        workbook,
        df_jam_lancar,
        df_jam_sibuk
    )

    # ---------------------------------------
    # Simpan Excel
    # ---------------------------------------
    workbook.save(OUTPUT_EXCEL)
    print("\nLaporan berhasil dibuat!")
    print(OUTPUT_EXCEL)
    
    # ==========================================================
# MENJALANKAN PROGRAM
# ==========================================================
if __name__ == "__main__":

    main()